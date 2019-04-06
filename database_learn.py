import pickle 
d = {}
integers = range(9999)
d["i"] = integers

f = open("08191.dat", "w+b")
pickle.dump(d, f) 
f.close()

f = open("08192.dat", "wb") # open or create the desired file in write-binary (wb) mode
pickle.dump(d, f, True) 
f.close()
# 比较两种编码格式的大小
import os 
s1 = os.stat("08191.dat").st_size # 得到两个文件的大小
s2 = os.stat("08192.dat").st_size
print("%d, %d, %.2f%%" % (s1, s2, (s2+0.0)/s1*100))

f = open("08191.dat", "rb") # open the desired file in read-binary (rb) mode
d = pickle.load(f)
print(d)
f.close()

"""三种类型的数据：
关系型数据库：MySQL、Microsoft Access、SQL Server、Oracle、...
非关系型数据库：MongoDB、BigTable(Google)、...
键值数据库：Apache Cassandra(Facebook)、LevelDB(Google) ..."""
# SQLite数据库的驱动已经在python里面了，所以，只要引用就可以直接使用 
import sqlite3 
conn = sqlite3.connect("0819.db") # 路径可以随意制定
cur = conn.cursor() # 建立游标对象 对数据库内容的操作，都是用游标对象方法来实现
create_table = "create table books (title text, author text, lang text)"
cur.execute(create_table) # 这样就在数据库0819.db中建立了一个表books，下面可以对这个表可以增加数据了。
cur.execute('insert into books values("diveintopython", "unknown", "chinese")') # 增加一组数据
# 为了保证数据能够保存,需执行以下操作
conn.commit()
cur.close()
conn.close()
# 查看存储的数据
conn = sqlite3.connect("0819.db")
cur = conn.cursor()
cur.execute('select * from books')
print(cur.fetchall())
# 添加多组数据
books_py = [("think python", "Allen Downey", "chinese"), ("Beginning pyhton", "Magnus Lie Hetland","english"), ("pythoncookbook", "熊熊", "chinese")]
cur.executemany('insert into books values (?, ?, ?)', books_py)
conn.commit()
books = cur.execute('select * from books')
for book in books:
    print(book)
# 更新数据中的信息
cur.execute("update books set title='深入python' where author = '熊熊'")
conn.commit()
cur.execute("select * from books where author = '熊熊'")
cur.fetchone()
# 删除数据
cur.execute("delete from books where author ='unknown'")
conn.commit()
cur.execute('select * from books')
cur.fetchall()
cur.close()
conn.close()

import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\iweut\Desktop\123.xlsx')
sheet = wb['Sheet1'] 
sheet.title
r = sheet.rows    # generator
c = sheet.columns
type(r)
type(c)
print(r, c)
len(list(r)) # 3
data = sheet['A1:C3']
print(data)
type(data)
import pprint 
pprint.pprint(data)
type(wb)
#wb.sheetnames()

sheet.title = "mydata"
sheet2 = wb.create_sheet(index=0, title="data2")
sheet2["B2"] = "tao"
wb.save(r'C:\Users\iweut\Desktop\123.xlsx') # 需要关闭文件才能保存修改
wb.remove(wb['data2'])
wb.save(r'C:\Users\iweut\Desktop\123.xlsx')